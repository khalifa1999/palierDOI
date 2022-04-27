import tkinter as tk
from tkinter import *
import tkinter.font
from tkinter import filedialog
import sys
import pandas as pd
from openpyxl import load_workbook

root = tk.Tk()
# make python game with the module
filename = 'C:/Users/diop9188/Desktop/tmp/test.xlsx'


# upload excel file
def select_excel():
    global filename
    filename = filedialog.asksaveasfilename(
        filetypes=(("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("Any file", "*")))
    calculus(filename)


# exit function
def get_out():
    sys.exit()


posa = 3
posb = 3

# places for corridors and non corridors
place = 2


def hors_corridors(worksheet):
    somme = 0
    global place
    loop = True
    while loop:
        if worksheet['J' + str(place)].value is not None:
            somme += worksheet['J' + str(place)].value
            print('somme' + str(somme))
        else:
            print('somme' + str(somme))
            worksheet['J' + str(place + 2)] = somme
            break
        place += 1


position = 2


def corridors(worksheet):
    somme = 0
    global position
    loop = True
    while loop:
        if worksheet['N' + str(position)].value is not None:
            somme += worksheet['N' + str(position)].value
        else:
            worksheet['N' + str(position + 1)] = somme
            break
        position += 1


def calculus(fic):
    global posa, posb
    wb = load_workbook(fic, data_only=True)
    ws = wb['Sheet']
    loop = True
    cpt = 0
    tab = 0
    while loop:

        if ws['K' + str(posb)].value is None and ws['J' + str(posa)].value is not None:
            print(ws['J' + str(posa)].value)
            cpt += ws['J' + str(posa)].value
            tab += 1
        else:
            sub = posa - tab - 1
            ws['L' + str(posa - 1)] = cpt + ws['J' + str(sub)].value
            print(ws['J' + str(sub)].value)
            print('value :' + str(ws['L' + str(posa - 1)]))
            cpt = 0
            tab = 0
            posb += 1
            posa += 1
            if ws['K' + str(posb)].value is None and ws['J' + str(posa)].value is not None:
                continue
            else:
                hors_corridors(ws)
                corridors(ws)
                wb.save(fic)
                sys.exit()

        # print('posa :'+str(posa))
        # print('posb :'+str(posb))
        posb += 1
        posa += 1


upload = Button(root, text='upload', command=select_excel, font=tkinter.font.Font(family="MS Shell Dlg 2", size=8),
                bg='green', fg='white', width=5)
quitter = Button(root, text='quitter', command=get_out, font=tkinter.font.Font(family="MS Shell Dlg 2", size=8),
                 bg='green', fg='white', width=5)
upload.grid(row=0, column=0)
quitter.grid(row=1, column=0)

root.mainloop()
