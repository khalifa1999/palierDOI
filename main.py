import tkinter as tk
from tkinter import *
from tkinter import filedialog
import sys
import pandas as pd
from openpyxl import load_workbook

root = tk.Tk()


# exit function
def get_out():
    sys.exit()


# titre root
# sf = "value is %s" % var.get()
root.title("Calcul balance")
filename = 'C:/Users/diop9188/Desktop/tmp/test.xlsx'


def select_excel():
    global filename
    filename = filedialog.asksaveasfilename(
        filetypes=(("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("Any file", "*")))


# ws = Worksheet()
# dictionnaire = ws.xlsx_to_dict(path=filename)
# print(dictionnaire)
# print(openpyxl.__version__)


df = pd.read_excel(filename, index_col=0)
df = df.where(pd.notnull(df), None)
dico = df.to_dict()['Deal']

# print(dico.values())
# For insertion into excel file
choix = ''


# This function is linked to a dropdown list
# Of operators and will be called everytime the user
def afficher(choice):
    global choix
    choix = var.get()
    global ven_default
    global volume_engage
    # On cree une variable global pour recuperer le volume reel en nombre
    global vr
    vr = vreel_default.get()
    dict_operateurs = dico
    # Modify here
    # For the Excel file
    for x, y in dict_operateurs.items():
        if choix == x:
            # ==> for testing purposes
            # print(y)
            volume_engage.destroy()
            ven_default = IntVar(root, value=y)
            volume_engage = Entry(root, textvariable=ven_default, bg="white", font=8, width=8)
            volume_engage.grid(row=6, column=1)
            calcul(vr, y)


posx = 2
posy = 2
posz = 2


def calcul(vr, ve):
    global fill1, fill2, fill3, Vfill1, Vfill2, Vfill3
    global result, vreel_default, Resultat
    global posx, posy, posz
    # Increment positions of our values when we access
    # the excel file to write the result

    # Initialisons les ca de chaque paliers à 0
    ca_p1 = 0
    ca_p2 = 0
    ca_p3 = 0
    # Volumes pallers
    v1 = Vfill1.get()
    v2 = Vfill2.get()
    v3 = Vfill3.get()
    # Les frais par palliers
    p1 = fill1.get()
    p2 = fill2.get()
    p3 = fill3.get()
    tarif_normal = tarif_normaldefault.get()
    result = 0
    result = vr * tarif_normal
    # basic if we have a deal and the max volume hasn't been reached
    basic = ve * tarif_normal
    # result = Resultat['text']
    # print(p1)
    # print(vr + "-" + result)

    if ve > vr > 0:
        # Excel insertion for regular price
        # In this condition, consumed volume is inferior to the engaged one
        diffa = basic - result
        eval1 = 'Deal. Situation de gain : +' + str(diffa)
        excel(choix, basic, eval1, posx, posy, posz)
        posx += 1
        posy += 1
        posz += 1
        Resultat.destroy()
        Resultat = Label(root, text=basic, bg='red')
        Resultat.grid(row=11, column=1)
    elif ve == 0:
        # Excel insertion for regular price
        # In this condition we didn't use a deal, we will just multiply the used volume to the regular rate
        eval2 = 'No deal'
        excel(choix, result, eval2, posx, posy, posz)
        posx += 1
        posy += 1
        posz += 1
        Resultat.destroy()
        Resultat = Label(root, text=result, bg='red')
        Resultat.grid(row=11, column=1)

    else:
        # This case is activate when the consumed volume is superior the regular one
        if vr > v1:
            # print(vr)
            ca_p1 = v1 * p1
            # print(ca_p1)
            vr -= v1
            # print(vr)
            if vr > v2:
                ca_p2 = v2 * p2
                vr -= v2
                print(vr)
                ca_p3 = vr * p3
                ca = ca_p1 + ca_p2 + ca_p3
                print(ca_p3)
                # tarif normal - deal
                tn = ve * tarif_normal
                diffc = ca - tn
                eval3 = 'Situation de perte : -' + str(diffc)
                excel(choix, ca, eval3, posx, posy, posz)
                posx += 1
                posy += 1
                posz += 1
                Resultat.destroy()
                Resultat = Label(root, text=ca, bg='red')
                Resultat.grid(row=11, column=1)


# CA-DEAL  display 0 , MODIFY LATER
# Rajouter evaluation
def excel(key, value, evaluation, posa, posb, posc):
    file_path = 'appending.xlsx'
    wb = load_workbook(file_path)
    ws = wb['Sheet']  # or wb.active
    ws['A1'] = 'Operateurs'
    ws['B1'] = 'CA-Deal'
    ws['C1'] = 'Evaluation'
    ws['A' + str(posa)] = key
    ws['B' + str(posb)] = value
    ws['C' + str(posc)] = evaluation
    wb.save(file_path)
    # worksheet.write('A' + str(posa), key)
    # worksheet.write('B' + str(posb), value)


var = StringVar(root)
choices = list(dico.keys())
# ['wind', 'lyca', 'free', 'libon', 'sima', 'lebara', 'USA retail', 'reptel', 'byt', 'ofr', 'us/canada', 'gap']
option = tk.OptionMenu(root, var, *choices, command=afficher)
var.set(choices[0])
option.grid(row=0, column=1)

# use command to call a function => select
# button = tk.Button(root, text="check value selected", command=select)


# text field
# default values variables
Vfill1 = IntVar(root, value=0)
Vfill2 = IntVar(root, value=0)
Vfill3 = IntVar(root, value=0)
fill1 = IntVar(root, value=0)
fill2 = IntVar(root, value=0)
fill3 = IntVar(root, value=0)
ven_default = IntVar(root, value=0)
vreel_default = IntVar(root, value=0)
tarif_normaldefault = IntVar(root, value=0)

# paliers volumes
VpalierI = tk.Label(root, text="Palier I-Volume", bg='orange', fg='white')
VpalierI.grid(row=1, column=0)

Vfill_palierI = tk.Entry(root, textvariable=Vfill1, bg="white", font=8, width=8)
Vfill_palierI.grid(row=2, column=0)

VpalierII = tk.Label(root, text="Palier II-Volume", bg='orange', fg='white')
VpalierII.grid(row=1, column=1)

Vfill_palierII = tk.Entry(root, textvariable=Vfill2, bg="white", font=8, width=8)
Vfill_palierII.grid(row=2, column=1)

VpalierIII = tk.Label(root, text="Palier III-Volume", bg='orange', fg='white')
VpalierIII.grid(row=1, column=2)

Vfill_palierIII = tk.Entry(root, textvariable=Vfill3, bg="white", font=8, width=8)
Vfill_palierIII.grid(row=2, column=2)

# paliers tarifs
palierI = tk.Label(root, text="Palier I-Tarif", bg='orange', fg='white')
palierI.grid(row=3, column=0)

fill_palierI = tk.Entry(root, textvariable=fill1, bg="white", font=8, width=8)
fill_palierI.grid(row=4, column=0)

palierII = tk.Label(root, text="Palier II-Tarif", bg='orange', fg='white')
palierII.grid(row=3, column=1)

fill_palierII = tk.Entry(root, textvariable=fill2, bg="white", font=8, width=8)
fill_palierII.grid(row=4, column=1)

palierIII = tk.Label(root, text="Palier III-Tarif", bg='orange', fg='white')
palierIII.grid(row=3, column=2)

fill_palierIII = tk.Entry(root, textvariable=fill3, bg="white", font=8, width=8)
fill_palierIII.grid(row=4, column=2)

# Volumes
label_vengage = Label(root, text='Volume Engage', bg='orange')
volume_engage = Entry(root, textvariable=ven_default, bg="white", font=8, width=8)

label_vengage.grid(row=5, column=1)
volume_engage.grid(row=6, column=1)

label_vreel = Label(root, text='Volume reel', bg='orange')
volume_reel = Entry(root, textvariable=vreel_default, bg="white", font=8, width=8)

label_vreel.grid(row=7, column=1)
volume_reel.grid(row=8, column=1)

label_tarifnormal = Label(root, text='Tarif normal', bg='orange')
tarif_normal = Entry(root, textvariable=tarif_normaldefault, bg='white', font=8, width=8)

label_tarifnormal.grid(row=9, column=1)
tarif_normal.grid(row=10, column=1)

# Label that where the result will be displayed
Resultat = Label(root, text="Chiffre d'affaire", bg='red')
Resultat.grid(row=11, column=1)

# buttons
upload = Button(root, text='upload', command=select_excel, bg='green', fg='white', font=6, width=5)
quitter = Button(root, text='quitter', command=get_out, bg='green', fg='white', font=6, width=5)
upload.grid(row=11, column=0)
quitter.grid(row=11, column=2)

root.mainloop()
