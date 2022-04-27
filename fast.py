import tkinter as tk
from tkinter import *
import tkinter.font
from tkinter import filedialog
import sys
import pandas as pd
from openpyxl import load_workbook

root = tk.Tk()
# make python game with the module
filename = 'C:/Users/diop9188/Desktop/tmp/test.xlsx'


# upload excel file
def select_excel():
    global filename
    filename = filedialog.asksaveasfilename(
        filetypes=(("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("Any file", "*")))
    calculus(filename)


# exit function
def get_out():
    sys.exit()


posa = 2


def calculus(fic):
    global posa
    wb = load_workbook(fic, data_only=True)
    ws = wb['Sheet']
    infinite = True
    ws['J1'] = 'Chiffre Affaire'
    cpt = 0

    while infinite:
        if ws['A' + str(posa)].value is not None:
            if ws['C' + str(posa)].value > ws['D' + str(posa)].value > 0:
                # ws['J' + str(posa)] = ws['C' + str(posa)].value * 2
                ws['J' + str(posa)] = ws['C' + str(posa)].value * ws['B' + str(posa)].value
                cpt += ws['J' + str(posa)].value
            elif ws['C' + str(posa)].value == 0:
                ws['J' + str(posa)] = ws['D' + str(posa)].value * ws['B' + str(posa)].value
                cpt += ws['J' + str(posa)].value
            else:
                if ws['D' + str(posa)].value > ws['F' + str(posa)].value:
                    p1 = ws['E' + str(posa)].value * ws['G' + str(posa)].value
                    p2 = (ws['F' + str(posa)].value - ws['E' + str(posa)].value) * ws['H' + str(posa)].value
                    p3 = (ws['D' + str(posa)].value - ws['F' + str(posa)].value) * ws['I' + str(posa)].value
                    ws['J' + str(posa)] = p1 + p2 + p3
                    cpt += ws['J' + str(posa)].value
                elif ws['F' + str(posa)].value > ws['D' + str(posa)].value > ws['E' + str(posa)].value:
                    p2 = (ws['D' + str(posa)].value - ws['E' + str(posa)].value) * ws['H' + str(posa)].value
                    p1 = ws['E' + str(posa)].value * ws['G' + str(posa)].value
                    ws['J' + str(posa)] = p1 + p2
                    cpt += ws['J' + str(posa)].value

        else:
            wb.save(fic)
            sys.exit()

        posa += 1


upload = Button(root, text='upload', command=select_excel, font=tkinter.font.Font(family="MS Shell Dlg 2", size=8),
                bg='green', fg='white', width=5)
quitter = Button(root, text='quitter', command=get_out, font=tkinter.font.Font(family="MS Shell Dlg 2", size=8),
                 bg='green', fg='white', width=5)
upload.grid(row=0, column=0)
quitter.grid(row=1, column=0)

root.mainloop()
